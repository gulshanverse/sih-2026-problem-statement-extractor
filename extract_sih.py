from __future__ import annotations

import argparse
import csv
import html
import re
import unicodedata
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

EXPECTED_RECORDS = 226

CANONICAL_COLUMNS = [
    "PS ID",
    "Problem Statement",
    "Description",
    "Organization",
    "Department",
    "Ministry",
    "Category",
    "Theme",
    "Type",
    "Technology",
    "Hardware/Software",
    "Submission Type",
    "Sector",
    "Location",
    "Reference Links",
    "Full Record Text",
    "Source Element",
]


def normalize_space(value: Any) -> str:
    if value is None:
        return ""
    text = str(value)
    text = html.unescape(text)
    text = text.replace("\xa0", " ")
    text = text.replace("\r", "")
    text = re.sub(r"\n\s*\n+", "\n\n", text)
    text = re.sub(r"[ \t]+", " ", text)
    return text.strip()


def clean_html_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value)
    if "<" in text and ">" in text:
        try:
            soup = BeautifulSoup(text, "html.parser")
            text = soup.get_text("\n", strip=False)
        except Exception:
            pass
    text = normalize_space(text)
    text = text.replace("\n ", "\n").replace(" \n", "\n")
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def sanitize_excel_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value)
    text = html.unescape(text)
    cleaned_chars = []
    for ch in text:
        if ch in {"\n", "\t", "\r"}:
            cleaned_chars.append(ch)
            continue
        cat = unicodedata.category(ch)
        if cat.startswith("C"):
            continue
        cleaned_chars.append(ch)
    text = "".join(cleaned_chars)
    text = text.replace("\uFFFD", "")
    if text.startswith("=") or text.startswith("+") or text.startswith("-") or text.startswith("@"):
        text = "'" + text
    if len(text) > 32767:
        text = text[:32767]
    return text


def safe_cell_value(value: Any) -> Any:
    if isinstance(value, str):
        return sanitize_excel_text(value)
    if value is None:
        return ""
    return value


def dedupe_key(record: Dict[str, str]) -> str:
    ps_id = normalize_space(record.get("PS ID", ""))
    title = normalize_space(record.get("Problem Statement", ""))
    description = normalize_space(record.get("Description", ""))
    return "|".join([ps_id.lower(), title.lower(), description.lower()])


def parse_modal_rows(modal_tag: Optional[Any]) -> Dict[str, str]:
    if modal_tag is None:
        return {}

    result: Dict[str, str] = {}
    table = modal_tag.select_one("table.table.table-bordered.table-hover")
    if not table:
        return result

    for row in table.select("tr"):
        cells = row.find_all(["th", "td"])
        if len(cells) < 2:
            continue
        key = clean_html_text(cells[0].get_text(" ", strip=True))
        value_cell = cells[1]
        value = ""
        if value_cell.find("div", class_="style-2"):
            value = clean_html_text(value_cell.find("div", class_="style-2").decode_contents())
        elif value_cell.find("a"):
            value = clean_html_text(value_cell.get_text(" ", strip=True))
        else:
            value = clean_html_text(value_cell.get_text(" ", strip=True))
        if key:
            result[key] = value

    return result


def parse_html_records(html_path: Path) -> List[Dict[str, str]]:
    html_text = html_path.read_text(encoding="utf-8", errors="ignore")
    soup = BeautifulSoup(html_text, "html.parser")
    rows = soup.select("#dataTablePS > tbody > tr")

    records: List[Dict[str, str]] = []
    seen: set[str] = set()

    for index, row in enumerate(rows, start=1):
        cells = row.find_all("td")
        if not cells:
            continue

        title_anchor = row.select_one("a[data-toggle='modal']") or row.select_one("a[data-target]")
        modal_id = None
        if title_anchor and title_anchor.has_attr("data-target"):
            modal_id = title_anchor.get("data-target", "").replace("#", "")

        modal_tag = soup.select_one(f"#{modal_id}") if modal_id else None
        modal_values = parse_modal_rows(modal_tag)

        organization = clean_html_text(cells[1].get_text(" ", strip=True)) if len(cells) > 1 else ""
        problem_statement = clean_html_text(title_anchor.get_text(" ", strip=True)) if title_anchor else clean_html_text(cells[2].get_text(" ", strip=True)) if len(cells) > 2 else ""
        category = clean_html_text(cells[3].get_text(" ", strip=True)) if len(cells) > 3 else ""
        ps_id = clean_html_text(cells[4].get_text(" ", strip=True)) if len(cells) > 4 else ""
        theme = clean_html_text(cells[6].get_text(" ", strip=True)) if len(cells) > 6 else ""

        modal_title = modal_values.get("Problem Statement Title") or problem_statement
        modal_description = modal_values.get("Description") or ""
        modal_org = modal_values.get("Organization") or organization
        modal_dept = modal_values.get("Department") or ""
        modal_category = modal_values.get("Category") or category
        modal_theme = modal_values.get("Theme") or theme
        modal_ministry = modal_values.get("Ministry") or ""
        modal_type = modal_values.get("Type") or ""
        modal_technology = modal_values.get("Technology") or ""
        modal_hw_sw = modal_values.get("Hardware/Software") or ""
        modal_submission = modal_values.get("Submission Type") or ""
        modal_sector = modal_values.get("Sector") or ""
        modal_location = modal_values.get("Location") or ""
        modal_refs = modal_values.get("Reference Links") or modal_values.get("Dataset Link") or modal_values.get("Youtube Link") or ""

        ps_id = modal_values.get("Problem Statement ID") or ps_id
        if ps_id and not str(ps_id).startswith("SIH"):
            ps_id = "SIH" + normalize_space(ps_id)
        ps_id = normalize_space(ps_id)

        description = clean_html_text(modal_description)
        if not description:
            description = clean_html_text(cells[2].get_text(" ", strip=True)) if len(cells) > 2 else ""

        full_record_text = clean_html_text(modal_tag.get_text("\n", strip=False)) if modal_tag else ""
        if not full_record_text:
            full_record_text = "\n".join([problem_statement, description, organization, category, theme, ps_id])

        record = {
            "PS ID": ps_id,
            "Problem Statement": clean_html_text(modal_title),
            "Description": description,
            "Organization": clean_html_text(modal_org),
            "Department": clean_html_text(modal_dept),
            "Ministry": clean_html_text(modal_ministry),
            "Category": clean_html_text(modal_category),
            "Theme": clean_html_text(modal_theme),
            "Type": clean_html_text(modal_type),
            "Technology": clean_html_text(modal_technology),
            "Hardware/Software": clean_html_text(modal_hw_sw),
            "Submission Type": clean_html_text(modal_submission),
            "Sector": clean_html_text(modal_sector),
            "Location": clean_html_text(modal_location),
            "Reference Links": clean_html_text(modal_refs),
            "Full Record Text": clean_html_text(full_record_text),
            "Source Element": f"row_{index}{f'|{modal_id}' if modal_id else ''}",
        }

        dedupe = dedupe_key(record)
        if not dedupe or dedupe in seen:
            continue
        seen.add(dedupe)
        records.append(record)

    return records


def build_summary_sheet(ws, total_count: int, dedup_count: int, source_file: str, timestamp: str) -> None:
    data = [
        ("Total Problem Statements", total_count),
        ("Expected Problem Statements", EXPECTED_RECORDS),
        ("Extraction Status", "PASS" if total_count == EXPECTED_RECORDS else "WARNING"),
        ("Duplicate Records Removed", dedup_count),
        ("Fields Extracted", len(CANONICAL_COLUMNS)),
        ("Source File", source_file),
        ("Extraction Timestamp", timestamp),
    ]
    ws.append(["Metric", "Value"])
    for key, value in data:
        ws.append([key, value])
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 26
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws["A1"].font = Font(bold=True)
    ws["B1"].font = Font(bold=True)
    for cell in ws[1]:
        cell.font = Font(bold=True)


def build_field_statistics(ws, records: List[Dict[str, str]]) -> None:
    ws.append(["Field", "Records Present"])
    for field in CANONICAL_COLUMNS:
        present = sum(1 for record in records if normalize_space(record.get(field, "")))
        ws.append([field, present])
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 16
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=False)
    for cell in ws[1]:
        cell.font = Font(bold=True)


def set_sheet_formatting(ws, header_row: int = 1, data_start_row: int = 2) -> None:
    ws.freeze_panes = f"A{data_start_row}"
    for cell in ws[header_row]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    for row_index in range(data_start_row, ws.max_row + 1):
        fill = PatternFill("solid", fgColor="FFFFFF" if row_index % 2 == 0 else "F7F9FC")
        for cell in ws[row_index]:
            cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col_index in range(1, ws.max_column + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=col_index, max_col=col_index):
            for cell in row:
                value = cell.value if cell.value is not None else ""
                if isinstance(value, str):
                    max_len = max(max_len, len(value))
                else:
                    max_len = max(max_len, len(str(value)))
        ws.column_dimensions[get_column_letter(col_index)].width = min(max(16, max_len + 2), 90)
    ws.auto_filter.ref = ws.dimensions


def export_xlsx(records: List[Dict[str, str]], output_path: Path, source_file: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Problem Statements"
    ws.append(CANONICAL_COLUMNS)
    for record in records:
        row = [safe_cell_value(record.get(field, "")) for field in CANONICAL_COLUMNS]
        ws.append(row)

    set_sheet_formatting(ws)

    summary_ws = wb.create_sheet("Summary")
    build_summary_sheet(summary_ws, len(records), 0, source_file, datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S %Z"))

    stats_ws = wb.create_sheet("Field Statistics")
    build_field_statistics(stats_ws, records)

    raw_ws = wb.create_sheet("Raw Data")
    raw_ws.append(CANONICAL_COLUMNS)
    for record in records:
        raw_ws.append([safe_cell_value(record.get(field, "")) for field in CANONICAL_COLUMNS])
    set_sheet_formatting(raw_ws)

    wb.save(output_path)


def export_csv(records: List[Dict[str, str]], csv_path: Path) -> None:
    with csv_path.open("w", encoding="utf-8", newline="") as fp:
        writer = csv.DictWriter(fp, fieldnames=CANONICAL_COLUMNS)
        writer.writeheader()
        for record in records:
            writer.writerow({key: safe_cell_value(record.get(key, "")) for key in CANONICAL_COLUMNS})


def validate_workbook(workbook_path: Path, records_count: int, expected_count: int = EXPECTED_RECORDS) -> Dict[str, Any]:
    if not workbook_path.exists():
        raise FileNotFoundError(f"Excel workbook not created: {workbook_path}")

    wb = load_workbook(workbook_path, read_only=False)
    required_sheets = {"Problem Statements", "Summary", "Field Statistics", "Raw Data"}
    actual_sheets = set(wb.sheetnames)
    missing = required_sheets - actual_sheets
    if missing:
        raise ValueError(f"Workbook missing required sheet(s): {sorted(missing)}")

    ws = wb["Problem Statements"]
    if ws.max_row < 2:
        raise ValueError("Problem Statements sheet has no records.")

    ps_ids = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0] is not None:
            ps_ids.append(str(row[0]).strip())

    duplicates = [ps_id for ps_id in ps_ids if ps_ids.count(ps_id) > 1]
    if duplicates:
        raise ValueError(f"Duplicate PS IDs detected: {sorted(set(duplicates))[:10]}")

    problem_statement_rows = [row for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2, values_only=True) if row[0] not in (None, "")]
    if not problem_statement_rows:
        raise ValueError("No problem statement titles were extracted.")

    max_text_length = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                if isinstance(cell.value, str):
                    max_text_length = max(max_text_length, len(cell.value))
    if max_text_length > 32767:
        raise ValueError(f"Workbook contains a cell text longer than Excel's limit: {max_text_length}")

    summary = {
        "records_found": records_count,
        "expected": expected_count,
        "duplicate_ps_ids": len(set(duplicates)),
        "sheets": sorted(wb.sheetnames),
        "status": "PASS" if records_count == expected_count else "WARNING",
    }
    return summary


def main() -> int:
    parser = argparse.ArgumentParser(description="Extract SIH 2026 problem statements from a local HTML dump.")
    parser.add_argument("--input", default="sih2026ps.html", help="Path to the HTML source file.")
    parser.add_argument("--output", default="output/SIH_2026_Problem_Statements.xlsx", help="Output XLSX path.")
    args = parser.parse_args()

    input_path = Path(args.input)
    if not input_path.exists():
        print(f"ERROR: Input file not found: {input_path}")
        return 1

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    records = parse_html_records(input_path)
    deduplicated = []
    seen = set()
    for record in records:
        key = dedupe_key(record)
        if key in seen:
            continue
        seen.add(key)
        deduplicated.append(record)

    csv_path = output_path.with_suffix(".csv")
    export_xlsx(deduplicated, output_path, str(input_path))
    export_csv(deduplicated, csv_path)

    validation = validate_workbook(output_path, len(deduplicated))
    print("========================================")
    print("SIH 2026 EXTRACTION REPORT")
    print("========================================")
    print("Input file:")
    print(input_path)
    print()
    print(f"Records found: {len(deduplicated)}")
    print(f"Expected: {EXPECTED_RECORDS}")
    print(f"Status: {validation['status']}")
    print(f"Columns: {len(CANONICAL_COLUMNS)}")
    print(f"Duplicate records removed: {max(0, len(records) - len(deduplicated))}")
    print(f"Records with PS ID: {sum(1 for r in deduplicated if normalize_space(r.get('PS ID', '')))}")
    print(f"Records with Description: {sum(1 for r in deduplicated if normalize_space(r.get('Description', '')))}")
    print(f"Records with Organization: {sum(1 for r in deduplicated if normalize_space(r.get('Organization', '')))}")
    print(f"Records with Theme: {sum(1 for r in deduplicated if normalize_space(r.get('Theme', '')))}")
    if len(deduplicated) != EXPECTED_RECORDS:
        print()
        print(f"WARNING: Expected {EXPECTED_RECORDS} records but found {len(deduplicated)}.")
    else:
        print()
        print("Excel generated successfully")
    print("========================================")

    return 0 if len(deduplicated) == EXPECTED_RECORDS else 2


if __name__ == "__main__":
    raise SystemExit(main())
